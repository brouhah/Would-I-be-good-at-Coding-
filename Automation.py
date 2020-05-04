#using an alias for openpyxl to make our code a bit cleaner
import openpyxl as xl
#module is used to import a Barchart and Reference
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
  #loads excel workbook and returns the workbook object
  wb = xl.load_workbook(filename)
  sheet = wb['Sheet1']

  #range function (1,4) would not include row 4
  for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

  values = Reference(
  sheet, 
  min_row=2, 
  max_row=sheet.max_row, 
  min_col=4, 
  max_col=4
  )

  chart = BarChart()
  chart.add_data(values)
  sheet.add_chart(chart, 'e2')

  wb.save(filename)
